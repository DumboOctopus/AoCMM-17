from openpyxl import load_workbook
import numpy
import math
from munkres import Munkres
import string

wb = load_workbook('Summary.xlsx') # the raw given

# processing the given and test quotes speed data into standard deviations for each person
givenSpeed = wb.get_sheet_by_name('Given Speed')

givenQuoteMean = []           # both count starting at 0, add one to indicie to get person number
givenQuoteStndDeviation = []

for x in range(11):
    temp = []
    for y in range(2, 10):
        temp.append(float(givenSpeed.cell(row=x+2, column=y).value))
    arr = numpy.array(temp)
    givenQuoteMean.append(numpy.mean(arr))
    givenQuoteStndDeviation.append(numpy.std(arr))

testValuesQuotes = []
testSpeed = wb.get_sheet_by_name('Test Speed')
testSpeed.cell(row=12, column=2).set_explicit_value(73.58)

for x in range(11):
    temp = []
    for y in range(2, 8):
        temp.append(float(testSpeed.cell(row=x+2, column=y).value))
    testValuesQuotes.append(temp)


# function for probability density
def probDensity(mean, stdDev, value):
    return (math.pow(math.e, -math.pow(value - mean, 2)/(2*math.pow(stdDev, 2)) / math.pow(2*math.pi*math.pow(stdDev, 2), 0.5)))



# on to the challenge of using Hungarian algorithm to minimize total error/cost (sum of values of (1-prob denstity)
def minCost(givenMean, givenStdDev, testValues):
    matrix = []

    for x in range(len(givenMean)):  # traversing base person
        temp = []
        matrix.append(temp)
        for y in range(len(testValues)):  # traversing person matched to
            sum = 0

            for i in range(len(testValues[y])):
                sum += 1 - probDensity(givenMean[x], givenStdDev[x], testValues[y][i])
            matrix[x].append(sum)

    m = Munkres()
    indicies = m.compute(matrix)

    print matrix[3][0]

    for row, column in indicies:
        print "(" + str(1 + row) + ", " + list(string.ascii_uppercase)[column] + ", " + str(matrix[row][column]) + ")"

minCost(givenQuoteMean, givenQuoteStndDeviation, testValuesQuotes)